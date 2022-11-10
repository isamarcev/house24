import os
from tempfile import NamedTemporaryFile

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Side, Border, Alignment, Font

from crm_accounting import models
from crm_home.models import Requisites
from home24.settings import BASE_DIR

from django.core.mail import send_mail, EmailMessage


def print_invoice(invoice, template_id, email=None):
    template = get_object_or_404(models.Template, id=template_id)
    data = {
        'total': invoice.amount,
        'invoiceNumber': invoice.number,
        'accountBalance': invoice.flat.personal_account.balance,
        'accountNumber': invoice.flat.personal_account.account_number,
        'invoiceMonth': f'{invoice.period_start.strftime("%d.%m")} '
                        f'- {invoice.period_end.strftime("%d.%m")}',
        'invoiceDate': invoice.date.strftime('%d.%m.%Y'),
        'invoiceAddress': f'{invoice.flat.owner} '
                          f'{invoice.flat.house.address} '
                          f'квартира №{invoice.flat.number}',
        'debt': invoice.flat.personal_account.balance - invoice.amount,
        'payCompany': f'{Requisites.objects.first()}'
    }

    template = load_workbook(template.file)
    sheet = template.active
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in data.keys():
                sheet[cell.coordinate] = data[cell.value]
    start_service = 19

    thin = Side(border_style="thin", color="000000")
    services = invoice.invoiceservice_set.all()
    for obj in services:
        sheet.insert_rows(start_service)

        sheet[f'A{start_service}'] = obj.service.name
        sheet[f'A{start_service}'].border = Border(top=thin, left=thin,
                                                   right=thin, bottom=thin)
        sheet[f'A{start_service}'].font = Font(size=12, italic=False,
                                               color="000000")
        sheet.merge_cells(f'A{start_service}:B{start_service}')

        sheet[f'C{start_service}'] = obj.invoice.tariff.name
        sheet[f'C{start_service}'].border = Border(top=thin, left=thin,
                                                   right=thin, bottom=thin)
        sheet[f'C{start_service}'].font = Font(size=12, italic=False,
                                               color="000000")
        sheet[f'C{start_service}'].alignment = Alignment(horizontal='right',
                                                         vertical='bottom')
        sheet.merge_cells(f'C{start_service}:D{start_service}')

        sheet[f'E{start_service}'] = obj.service.unit.title
        sheet[f'E{start_service}'].border = Border(top=thin, left=thin,
                                                   right=thin, bottom=thin)
        sheet[f'E{start_service}'].font = Font(size=12, italic=False,
                                               color="000000")
        sheet[f'E{start_service}'].alignment = Alignment(horizontal='right',
                                                         vertical='bottom')
        sheet.merge_cells(f'E{start_service}:F{start_service}')

        sheet[f'G{start_service}'] = obj.amount
        sheet[f'G{start_service}'].border = Border(top=thin, left=thin,
                                                   right=thin, bottom=thin)
        sheet[f'G{start_service}'].font = Font(size=12, italic=False,
                                               color="000000")
        sheet[f'G{start_service}'].alignment = Alignment(horizontal='right',
                                                         vertical='bottom')
        sheet.merge_cells(f'G{start_service}:H{start_service}')

        sheet[f'I{start_service}'] = obj.total
        sheet[f'I{start_service}'].border = Border(top=thin, left=thin,
                                                   right=thin, bottom=thin)
        sheet[f'I{start_service}'].font = Font(size=12, italic=False,
                                               color="000000")
        sheet[f'I{start_service}'].alignment = Alignment(horizontal='right',
                                                         vertical='bottom')
        sheet.merge_cells(f'I{start_service}:K{start_service}')

        start_service += 1

    sheet.merge_cells(f'A{start_service}:B{start_service}')
    sheet.merge_cells(f'C{start_service}:D{start_service}')
    sheet.merge_cells(f'E{start_service}:F{start_service}')
    sheet.merge_cells(f'G{start_service}:H{start_service}')
    sheet.merge_cells(f'I{start_service}:K{start_service}')
    with NamedTemporaryFile(prefix='.xlsx') as tmp:
        temp_path = tmp.name
        template.save(temp_path)
        tmp.seek(0)
        stream = tmp.read()
    filepath = os.path.join(BASE_DIR, temp_path)
    if email:
        email_address = invoice.flat.owner.email
        if email_address:
            email = EmailMessage(
                'CRM 24 Administarions',
                f'Квитанция',
                None,
                [email_address]
            )
            email.attach(f'Квитанция № {invoice.number}.xlsx', stream,
                         'application/vnd.ms-excel')
            email.send()
            return True
        else:
            return False
    response = HttpResponse(stream, content_type=filepath)
    response['Content-Disposition'] \
        = "attachment; filename=%s" % f'{invoice.number}.xlsx'
    return response