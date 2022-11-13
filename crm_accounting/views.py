import mimetypes
import os
from tempfile import NamedTemporaryFile

from django.contrib import messages
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin
from django.db.models import Q
from django.db.models.aggregates import Sum
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, DetailView, \
    ListView, DeleteView, View, TemplateView
from django_datatables_view.base_datatable_view import BaseDatatableView
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Side, Border, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from crm_home.models import PaymentState, TariffService, Requisites
from home24.settings import BASE_DIR
from houses.models import Section, Flat, House
from users.models import CustomUser
from . import models
from . import forms
from .services import xls_invoices
from .services.xls_invoices import print_invoice


class AdminPermissionMixin(LoginRequiredMixin, UserPassesTestMixin):
    login_url = reverse_lazy('users:admin-login')
    redirect_field_name = 'redirect_to'

    def test_func(self):
        if self.request.user.role:
            return getattr(self.request.user.role,
                           str(self.check_permission_name))

    def handle_no_permission(self):
        if self.request.user.is_authenticated and not self.request.user.role:
            return HttpResponseRedirect(reverse_lazy('users:user_profile'))
        return super().handle_no_permission()

def main_page(request, pk):
    pass


def get_statistics(context):
    """Balance personal accounts, depts and cahsbox`s state """
    try:
        income = models.Transaction.objects. \
            filter(payment_state__type='in').aggregate(Sum('amount'))
        outcome = models.Transaction.objects. \
            filter(payment_state__type='out').aggregate(Sum('amount'))
        cashbox_state = income['amount__sum'] - outcome['amount__sum']
        context['cashbox_state'] = cashbox_state
    except TypeError:
        context['cashbox_state'] = 0
    sum_personal_accounts = models.PersonalAccount.objects.all().aggregate(
        Sum('balance'))
    if sum_personal_accounts:
        context['sum_accounts_balance'] = sum_personal_accounts.\
            get('balance__sum')
    invoices_dopts = models.Invoice.objects.filter(
        payment_state=True,status__in=['Неоплачена', 'Частично оплачена']). \
        aggregate(Sum('amount'))
    if invoices_dopts:
        context['invoices_dobts'] = invoices_dopts.get('amount__sum')
    return context


class AccountCreateView(AdminPermissionMixin, CreateView):
    model = models.PersonalAccount
    form_class = forms.PersonalAccountForm
    success_url = reverse_lazy('crm_accounting:account_list')
    template_name = 'crm_accounting/personalaccount_form.html'
    check_permission_name = 'personal_account'

    def post(self, request, *args, **kwargs):
        form_class = self.form_class(request.POST or None)
        if form_class.is_valid():
            flat = request.POST.get('flat')
            form_class.save()
            if flat:
                flat = Flat.objects.get(pk=flat)
                flat.personal_account = form_class.instance
            return HttpResponseRedirect(self.success_url)
        else:
            return render(request, self.template_name,
                          context={'form': form_class})


class PersonalAccountUpdateView(AdminPermissionMixin, UpdateView):
    model = models.PersonalAccount
    template_name = 'crm_accounting/personalaccount_update_form.html'
    form_class = forms.PersonalAccountForm
    success_url = reverse_lazy('crm_accounting:account_list')
    check_permission_name = 'personal_account'

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        account = self.get_object()
        context['sections'] = Section.objects.filter(house=account.house)
        context['flats'] = Flat.objects.filter(section=account.section). \
            order_by('number')
        return context

    def post(self, request, *args, **kwargs):
        instance = self.get_object()
        form_class = self.form_class(request.POST or None, instance=instance)
        if form_class.is_valid():
            flat = request.POST.get('flat')
            form_class.save()
            if flat:
                flat = Flat.objects.get(pk=flat)
                flat.personal_account = form_class.instance
            return HttpResponseRedirect(self.success_url)
        else:
            return render(request, self.template_name,
                          context={'form': form_class})


class AccountListView(AdminPermissionMixin, ListView):
    model = models.PersonalAccount
    queryset = None
    check_permission_name = 'personal_account'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = dict()
        context['houses'] = House.objects.all()
        context['owner'] = CustomUser.objects.all()
        context = get_statistics(context)
        return context


class AccountListViewAjax(BaseDatatableView):
    model = models.PersonalAccount
    columns = ['account_number', 'status', 'flat.number',
               'house', 'section', 'flat.owner', 'balance', 'id']
    order_columns = ['account_number', 'status', 'flat.number',
                     'house', 'section', 'flat.owner', 'balance', 'id']

    def get_initial_queryset(self):
        return self.model.objects.all().select_related('house', 'section',
                                                       'flat__owner')

    def filter_queryset(self, qs):
        number = self.request.GET.get('number')
        house = self.request.GET.get('house')
        section = self.request.GET.get('section')
        flat = self.request.GET.get('flat')
        status = self.request.GET.get('status')
        owner = self.request.GET.get('owner')
        dolg = self.request.GET.get('dolg')
        if number:
            qs = qs.filter(account_number__icontains=number)
        if house:
            qs = qs.filter(house=house)
        if section:
            qs = qs.filter(section=section)
        if flat:
            qs = qs.filter(flat__number=flat)
        if owner:
            qs = qs.filter(flat__owner=owner)
        if dolg:
            if dolg == 'false':
                qs = qs.filter(balance__gte=0)
            elif dolg == 'true':
                qs = qs.filter(balance__lt=0)
        if status:
            qs = qs.filter(status=status)
        return qs


class DeletePersonalAccount(DeleteView):
    model = models.PersonalAccount
    check_permission_name = 'personal_account'

    def post(self, request, *args, **kwargs):
        personal_account = self.model.objects.get(
            pk=request.POST.get('account'))
        if request.user.is_superuser:
            personal_account.delete()
            return JsonResponse({'success': 'success'})
        else:
            return JsonResponse(
                {'success': 'Удаление счета доступно только Директору'})


class DownloadExcelAccounts(View):

    def get(self, request):
        number = self.request.GET.get('number')
        status = self.request.GET.get('status')
        flat = self.request.GET.get('flat')
        house = self.request.GET.get('house')
        section = self.request.GET.get('section')
        owner = self.request.GET.get('owner')
        dolg = self.request.GET.get('dolg')
        qs = models.PersonalAccount.objects.all()
        if number:
            qs = qs.filter(account_number__icontains=number)
        if status:
            qs = qs.filter(status=status)
        if flat:
            qs = qs.filter(flat__number=flat)
        if house:
            qs = qs.filter(house_id=house)
        if section:
            qs = qs.filter(flat__section_id=section)
        if owner:
            qs = qs.filter(flat__owner_id=owner)
        if dolg:
            if dolg == 'true':
                qs = qs.filter(balance__lt=0)
            else:
                qs = qs.filter(balance__gte=0)
        if qs:
            wb = Workbook()
            sheet = wb.active
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 30
            sheet.column_dimensions['G'].width = 10
            sheet.title = 'rabge names'
            sheet['A1'] = 'Лицевой счет'
            sheet['B1'] = 'Статус'
            sheet['C1'] = 'Дом'
            sheet['D1'] = 'Секция'
            sheet['E1'] = 'Квартира'
            sheet['F1'] = 'Владелец'
            sheet['G1'] = 'Остаток'
            len_qs = len(qs) + 2
            start_row = 2
            while start_row != len_qs:
                for obj in qs:
                    sheet.insert_rows(start_row)
                    sheet[f'A{start_row}'] = obj.account_number
                    sheet[f'B{start_row}'] = obj.status
                    sheet[f'C{start_row}'] = f'{obj.house}'
                    sheet[f'D{start_row}'] = f'{obj.section}'
                    sheet[f'E{start_row}'] = f'{obj.flat}'
                    sheet[f'F{start_row}'] = f'{obj.flat.owner}'
                    sheet[f'G{start_row}'] = f'{obj.balance}'
                    start_row += 1
            with NamedTemporaryFile(prefix='.xlsx') as tmp:
                temp_path = tmp.name
                wb.save(temp_path)
                tmp.seek(0)
                stream = tmp.read()
            filepath = os.path.join(BASE_DIR, temp_path)
            response = HttpResponse(stream, content_type=filepath)
            date = datetime.today().strftime("%Y%m%d")
            response['Content-Disposition'] \
                = "attachment; filename=%s" % f'accounts_{date}.xlsx'
            return response


class PersonalAccountDetailView(AdminPermissionMixin, DetailView):
    model = models.PersonalAccount
    template_name = 'crm_accounting/personalaccount_detail.html'
    check_permission_name = 'personal_account'


def get_flats(request):
    if request.GET:
        flat_list = list()
        if request.GET.get('section'):
            section = request.GET.get('section')
            flats = Flat.objects.filter(section=section). \
                select_related('owner')
            for flat in flats:
                element = {
                    'flat_id': flat.id,
                    'flat_number': flat.number,
                }
                flat_list.append(element)
        return JsonResponse({'flats': flat_list}, status=200)


def get_users(request):
    if request.GET:
        owner_info = dict()
        if request.GET.get('flat'):
            flat = request.GET.get('flat')
            try:
                owner = CustomUser.objects.get(flat=flat)
                owner_info['name'] = \
                    f'{owner.first_name} {owner.last_name} {owner.father_name}'
                owner_info['phone'] = owner.phone
                owner_info['id'] = owner.id
            finally:
                pass
        return JsonResponse({'owner': owner_info}, status=200)


def calculate_balance(form_class, **kwargs):
    """Check the form Transaction for checkbox and calculate account balance"""
    form_class.save(commit=False)
    form_type = kwargs.get('type')
    payment_state = form_class.instance.payment_state.type

    if form_type == 'update' and payment_state == 'in':
        completed_instance = kwargs.get('completed')
        completed_instance_form = form_class.instance.completed
        account = models.PersonalAccount.objects.get(
            id=form_class.instance.personal_account.id)
        if completed_instance == completed_instance_form:
            pass
        elif completed_instance_form:
            operation = account.balance + form_class.instance.amount
            account.balance = operation
            account.save()
        elif not completed_instance_form:
            operation = account.balance - form_class.instance.amount
            account.balance = operation
            account.save()
    elif form_type == 'create' and payment_state == 'in':
        account = models.PersonalAccount.objects.get(
            id=form_class.instance.personal_account.id)
        if form_class.instance.completed:
            operation = account.balance + form_class.instance.amount
            account.balance = operation
            account.save()
    form_class.save()


class TransactionCreateView(AdminPermissionMixin, CreateView):
    model = models.Transaction
    form_class = forms.TransactionForm
    success_url = reverse_lazy('crm_accounting:transaction_list')
    check_permission_name = 'cashbox'

    def get_context_data(self, **kwargs):
        context = dict()
        context['manager_list'] = CustomUser.objects.filter(~Q(role=None)). \
            select_related('role')
        type_of_transaction = self.request.GET.get('type')
        context['manager'] = self.request.user
        instance_id = self.request.GET.get('transaction_id')
        flat_id = self.request.GET.get('flat_id')

        if kwargs.get('form'):
            context['form'] = kwargs['form']
        else:
            context['form'] = self.form_class()
        if instance_id:
            transaction = models.Transaction.objects.select_related('manager',
                                                                    'owner'). \
                get(pk=instance_id)
            initial = {
                'payment_state': transaction.payment_state,
                'comment': transaction.comment,
                'owner': transaction.owner,
                'personal_account': transaction.personal_account,
                'amount': transaction.amount,
                'manager': transaction.manager,
                'completed': transaction.completed
            }
            context['form'] = self.form_class(initial=initial)
            type_of_transaction = transaction.payment_state.type
            context['manager'] = transaction.manager
        if flat_id:
            flat = get_object_or_404(Flat, id=flat_id)
            initial = {
                'owner': flat.owner,
                'personal_account': flat.personal_account,

            }
            context['form'] = self.form_class(initial=initial)
        if type_of_transaction == 'in':
            context['type'] = 'приходная ведомость'
        elif type_of_transaction == 'out':
            context['type'] = 'расходная ведомость'
        context['payment_states'] = PaymentState.objects.filter(
            type=type_of_transaction)
        print(context)
        print(self.form_class)
        return context

    def post(self, request, *args, **kwargs):
        form_class = self.form_class(request.POST or None)
        if form_class.is_valid():
            calculate_balance(form_class, type='create')
            return HttpResponseRedirect(self.success_url)
        else:
            return render(request, 'crm_accounting/transaction_form.html',
                          context={'form': form_class})


def get_personal_accounts_ajax(request):
    owner_id = request.GET.get('owner_id')
    accounts = models.PersonalAccount.objects.filter(
        account_flat__owner_id=owner_id)
    account_list = list()
    for i in accounts:
        account = {
            'id': i.id,
            'account_number': i.account_number
        }
        account_list.append(account)
    return JsonResponse({'accountList': account_list})


class TransactionListView(AdminPermissionMixin, ListView):
    model = models.Transaction
    queryset = None
    check_permission_name = 'cashbox'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = dict()
        context['owners'] = CustomUser.objects.filter(role=None)
        context['payment_states'] = PaymentState.objects.all()
        account_id = self.request.GET.get('account_id')
        if account_id:
            account_id = get_object_or_404(models.PersonalAccount,
                                           id=account_id)
            context['account_id'] = account_id.account_number
            print(account_id.account_number)

        context = get_statistics(context)
        return context


def filter_qs_daterange(date, qs):
    date_start = datetime.strptime(date.split(' - ')[0],
                                   '%m-%d-%Y')
    date_end = datetime.strptime(date.split(' - ')[1],
                                 '%m-%d-%Y')
    qs = qs.filter(Q(date__gt=date_start), Q(date__lt=date_end))
    return qs


class TransactionListViewAjax(BaseDatatableView):
    model = models.Transaction
    columns = ['number', 'date', 'completed', 'payment_state',
               'owner', 'personal_account', 'payment_state.type',
               'amount', 'id']
    order_columns = ['date']

    def get_initial_queryset(self):
        if self.request.GET.get('account'):
            queryset = self.model.objects.filter(
                personal_account_id=self.request.GET.get('account'))
        else:
            queryset = self.model.objects.all()
        return queryset.select_related('payment_state'). \
            order_by('-number')

    def filter_queryset(self, qs):
        number = self.request.GET.get('number')
        date = self.request.GET.get('date')
        owner = self.request.GET.get('owner')
        personal_account = self.request.GET.get('personal_account')
        completed = self.request.GET.get('completed')
        payment_state = self.request.GET.get('paymentstate')
        type_of_payment = self.request.GET.get('type_of_payment')
        if type_of_payment:
            qs = qs.filter(payment_state__type=type_of_payment)
        if number:
            qs = qs.filter(number__contains=number)
        if owner:
            qs = qs.filter(owner=owner)
        if personal_account:
            qs = qs.filter(
                personal_account__account_number__contains=personal_account)
        if completed:
            if completed == 'completed':
                qs = qs.filter(completed=True)
            elif completed == 'not_completed':
                qs = qs.filter(completed=False)
        if payment_state:
            qs = qs.filter(payment_state=payment_state)
        if date:
            qs = filter_qs_daterange(date, qs)
        return qs


class DeleteTransaction(DeleteView):
    model = models.Transaction


    def get(self, request, *args, **kwargs):
        transaction = self.model.objects.get(
            pk=request.GET.get('transaction')
        )
        transaction.delete()
        return HttpResponseRedirect(
            reverse_lazy('crm_accounting:transaction_list'))

    def post(self, request, *args, **kwargs):
        transaction = self.model.objects.get(
            pk=request.POST.get('account'))
        if request.user.is_superuser:
            transaction.delete()
            return JsonResponse({'success': 'success'})
        else:
            return JsonResponse({'success': 'Удаление счета доступно '
                                            'только старшему персоналу!'})


class TransactionUpdateView(AdminPermissionMixin, UpdateView):
    model = models.Transaction
    form_class = forms.TransactionForm
    success_url = reverse_lazy('crm_accounting:transaction_list')
    template_name = 'crm_accounting/transaction_update_form.html'
    check_permission_name = 'cashbox'


    def get_context_data(self, **kwargs):
        instance = self.get_object()
        context = dict()
        if kwargs.get('form'):
            context['form'] = kwargs['form']
        else:
            context['form'] = self.form_class(instance=instance)
        context['object'] = instance
        instance = context['form'].instance
        type_of_transaction = instance.payment_state.type
        if type_of_transaction == 'in':
            context['type'] = 'Приходная ведомость'
        elif type_of_transaction == 'out':
            context['type'] = 'Расходная ведомость'
        context['payment_states'] = PaymentState.objects.filter(
            type=type_of_transaction)
        context['manager_list'] = CustomUser.objects.filter(~Q(role=None)). \
            select_related('role')
        return context

    def post(self, request, *args, **kwargs):
        instance = self.get_object()
        completed_instance = instance.completed
        form_class = self.form_class(request.POST or None, instance=instance)
        if form_class.is_valid():
            calculate_balance(form_class, type='update',
                              completed=completed_instance)
            return HttpResponseRedirect(self.success_url)
        else:
            return render(request,
                          'crm_accounting/transaction_update_form.html',
                          context={'form': form_class})


class TransactionDetailView(AdminPermissionMixin, DetailView):
    model = models.Transaction
    check_permission_name = 'cashbox'

    def get_queryset(self):
        return self.model.objects.all(). \
            select_related('personal_account__flat__owner', 'payment_state')

    def get_context_data(self, **kwargs):
        context = super(TransactionDetailView, self).get_context_data()
        if self.object.payment_state.type == 'in':
            context['type'] = 'Приходная ведомость'
        else:
            context['type'] = 'Расходная ведомость'
        return context


def download_file(request):
    transaction_id = request.GET.get('transaction_id')
    transaction = get_object_or_404(models.Transaction.objects.
                                    select_related('payment_state',
                                                   'owner',
                                                   'manager'),
                                    id=transaction_id)
    wb = Workbook()
    ws = wb.active
    ws.title = 'rabge names'
    ws['A1'] = 'Платеж'
    ws['A2'] = 'Дата'
    ws['A3'] = 'Владелец квартиры'
    ws['A4'] = 'Лицевой счет'
    ws['A5'] = 'Приход/расход'
    ws['A6'] = 'Статус'
    ws['A7'] = 'Статья'
    ws['A8'] = 'Квитанция'
    ws['A9'] = 'Услуга'
    ws['A10'] = 'Сумма'
    ws['A11'] = 'Валюта'
    ws['A12'] = 'Комментарий'
    ws['A13'] = 'Менеджер'
    ws['B1'] = f'#{transaction.number}'
    ws['B2'] = f'{transaction.date}'
    if transaction.owner:
        ws['B3'] = f'{transaction.owner}'
    if transaction.personal_account:
        ws['B4'] = f'{transaction.personal_account}'
    if transaction.payment_state.type == 'in':
        ws['B5'] = f'Приход'
    else:
        ws['B5'] = f'Расход'
    if transaction.completed:
        ws['B6'] = "Проведен"
    else:
        ws['B6'] = 'Не проведен'
    ws['B7'] = f'{transaction.payment_state.title}'
    ws['B8'] = ''
    ws['B9'] = ''
    ws['B10'] = transaction.amount
    ws['B11'] = 'UAH'
    ws['B12'] = transaction.comment
    ws['B13'] = f'{transaction.manager}'
    with NamedTemporaryFile(prefix='.xlsx') as tmp:
        temp_path = tmp.name
        wb.save(temp_path)
        tmp.seek(0)
        stream = tmp.read()
    filepath = os.path.join(BASE_DIR, temp_path)
    response = HttpResponse(stream, content_type=filepath)
    response['Content-Disposition'] \
        = "attachment; filename=%s" % f'{transaction.number}.xlsx'
    return response


class InvoiceListView(AdminPermissionMixin, ListView):
    model = models.Invoice
    check_permission_name = 'invoice'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(InvoiceListView, self).get_context_data()
        context['owner'] = CustomUser.objects.filter(role=None)
        flat_id = self.request.GET.get('flat_id')
        if flat_id:
            flat = get_object_or_404(Flat, id=flat_id)
            context['flat_number'] = flat.number
        context = get_statistics(context)
        return context


class InvoiceListViewAjax(BaseDatatableView):
    model = models.Invoice
    columns = ['number', 'status', 'date', 'flat',
               'house', 'flat.owner', 'payment_state',
               'amount', 'id']
    order_columns = ['date']

    def get_initial_queryset(self):
        return self.model.objects.all().select_related('house', 'section',
                                                       'flat__owner').\
            order_by('-id')

    def filter_queryset(self, qs):
        number = self.request.GET.get('number')
        status = self.request.GET.get('status')
        date = self.request.GET.get('date')
        month = self.request.GET.get('month')
        month_list = [
            'January', 'February', 'March', 'April', 'May', 'June', 'July',
            'August', 'September', 'October', 'November', 'December'
        ]
        if month:
            filter_month = month.split(' ')[0]
            qs = qs.filter(date__month=int(month_list.index(filter_month) + 1))
        if date:
            qs = filter_qs_daterange(date, qs)
        flat = self.request.GET.get('flat')
        owner = self.request.GET.get('owner')
        payment_state = self.request.GET.get('payment_state')
        if number:
            qs = qs.filter(account_number__icontains=number)
        if status:
            qs = qs.filter(status=status)
        if owner:
            qs = qs.filter(flat__owner_id=owner)
        if flat:
            qs = qs.filter(flat__number=flat)
        if payment_state:
            qs = qs.filter(payment_state=payment_state)
        return qs


class InvoiceCreateView(AdminPermissionMixin, CreateView):
    model = models.Invoice
    service_formset = forms.InvoiceServiceFormset
    form_class = forms.InvoiceForm
    check_permission_name = 'invoice'

    def get_context_data(self, **kwargs):
        context = super(InvoiceCreateView, self).get_context_data()
        personal_account_id = self.request.GET.get('personal_account')
        if personal_account_id:
            personal_account = get_object_or_404(
                models.PersonalAccount.objects.
                select_related('flat__section__house', 'flat__owner'),
                id=personal_account_id)
            initial = {
                'personal_account': personal_account.account_number,
                'flat': personal_account.flat,
                'house': personal_account.flat.house,
                'section': personal_account.flat.section,
            }
            context['form'] = self.form_class(initial=initial)
        date_service = {'form-TOTAL_FORMS': '0',
                        'form-INITIAL_FORMS': '0',
                        }
        context['service_formset'] = self.service_formset(data=date_service)
        invoice_id = self.request.GET.get('invoice_id')
        if invoice_id:
            invoice_for_copy = get_object_or_404(models.Invoice, id=invoice_id)
            initial = {
                'flat': invoice_for_copy.flat,
                'house': invoice_for_copy.flat.house,
                'section': invoice_for_copy.flat.section,
                'tariff': invoice_for_copy.tariff,
                'personal_account': invoice_for_copy.flat.personal_account,
                'amount': invoice_for_copy.amount,
                'status': invoice_for_copy.status,
                'payment_state': invoice_for_copy.payment_state
            }
            context['form'] = self.form_class(initial=initial)
            context['service_formset'] = self.service_formset(
                queryset=models.InvoiceService.objects.filter(
                    invoice=invoice_for_copy).select_related('service__unit'))
        context['services'] = forms.Service.objects.all()
        context['units'] = forms.Unit.objects.all()
        return context

    def post(self, request, *args, **kwargs):
        form_class = self.form_class(request.POST or None)
        service_formset = self.service_formset(request.POST or None)
        if all([form_class.is_valid(), service_formset.is_valid()]):
            form_class.save()
            service_formset.save(commit=False)
            for form in service_formset:
                if form.instance not in service_formset.deleted_objects\
                        and all([form.instance.service, form.instance.total]):
                    form.instance.pk = None
                    form.instance._state.adding = True
                    form.instance.invoice = form_class.instance
                    form.instance.save()
            form_class.calculate_invoice(form_class)
            return HttpResponseRedirect(reverse_lazy
                                        ('crm_accounting:invoice_list'))
        return render(request, 'crm_accounting/invoice_form.html',
                      context={'form':form_class,
                               'service_formset':service_formset})


class InvoiceDetailView(AdminPermissionMixin, DetailView):
    model = models.Invoice
    check_permission_name = 'invoice'

    def get_queryset(self):
        queryset = models.Invoice.objects.all().\
            select_related('flat__personal_account', 'flat__owner', ).\
            prefetch_related('invoiceservice_set__service__unit')
        return queryset


class InvoiceUpdateView(AdminPermissionMixin, UpdateView):
    model = models.Invoice
    template_name = 'crm_accounting/invoice_update_form.html'
    service_formset = forms.InvoiceServiceFormset
    form_class = forms.InvoiceForm
    queryset = model.objects.all().select_related('flat__section__house')
    check_permission_name = 'invoice'

    def get_context_data(self, **kwargs):
        context = dict()
        if kwargs.get('form'):
            context['form'] = kwargs['form']
            self.object = kwargs['form'].instance
        else:
            context['form'] = self.form_class(instance=self.object)
        context['object'] = self.object
        context['sections'] = Section.objects.filter(house=self.object.house)
        context['flats'] = Flat.objects.filter(
            section_id=self.object.section.id)
        context['service_formset'] = self.service_formset(
            queryset=models.InvoiceService.objects.filter(
                invoice=self.object))
        context['services'] = forms.Service.objects.all()
        context['units'] = forms.Unit.objects.all()
        return context

    def post(self, request, *args, **kwargs):
        instance = self.get_object()
        form_class = self.form_class(request.POST or None,
                                     instance=instance)
        service_formset = self.service_formset(
            request.POST,
            queryset=models.InvoiceService.objects.filter(invoice=instance))
        if all([form_class.is_valid(), service_formset.is_valid()]):
            form_class.save()
            service_formset.save(commit=False)
            for form in service_formset:
                if form.instance not in service_formset.deleted_objects \
                        and all([form.instance.service, form.instance.total]):
                    form.instance.invoice = form_class.instance
                    form.instance.save()
            for form in service_formset.deleted_objects:
                form.delete()
            form_class.calculate_invoice(form_class)
            messages.success(request, "Квитанция изменена")
            return HttpResponseRedirect(reverse_lazy
                                        ('crm_accounting:invoice_list'))
        return render(request, self.template_name,
                      self.get_context_data(form=form_class,
                                            service_formset=service_formset))


def calculate_invoice(account_id):
    """Calculate balance of PERSONAL ACCOUNT"""
    account = models.PersonalAccount.objects.get(id=account_id)
    income_balance = models.Transaction.objects.filter(
        personal_account=account,
        payment_state__type='in',
        completed=True).\
        aggregate(Sum('amount')).get('amount__sum')
    outcome_balance = models.Invoice.objects.filter(
        personal_account=account,
        payment_state=True,
        status='Оплачена').\
        aggregate(Sum('amount')).get('amount__sum')
    if not income_balance:
        income_balance = 0
    if not outcome_balance:
        outcome_balance = 0
    account.balance = income_balance - outcome_balance
    account.save()


class InvoiceDeleteView(DeleteView):
    model = models.Invoice

    def post(self, request, *args, **kwargs):
        delete_list = request.POST.get('deleted_list').split(',')
        if request.user.is_superuser:
            invoices = self.model.objects.filter(id__in=delete_list)
            for invoice in invoices:
                calculate_invoice(invoice.personal_account)
                invoice.delete()
            return JsonResponse({'success': 'success'})
        else:
            return JsonResponse({'success': 'У Вас нет прав для удаления!'})


class TemplatesUpdateView(TemplateView):
    template_name = 'crm_accounting/template_update_view.html'

    def get_context_data(self, **kwargs):
        context = dict()
        context['templates'] = models.Template.objects.all()
        context['form_template'] = forms.TemplateForm
        return context

    def get(self, request):
        delete_id = request.GET.get('delete_id')
        default_id = request.GET.get('default_id')
        if delete_id:
            models.Template.objects.filter(id=delete_id).delete()
        if default_id:
            templates = models.Template.objects.all()
            for template in templates:
                if template.id == int(default_id):
                    template.default = True
                else:
                    template.default = False
                template.save()
        return super().get(self)

    def post(self, request):
        form_template = forms.TemplateForm(request.POST, request.FILES)
        if form_template.is_valid():
            if form_template.cleaned_data.get('file'):
                form_template.save()
            return HttpResponseRedirect(
                reverse_lazy('crm_accounting:invoice_list'))
        else:
            return render(request, self.template_name, context={
                'templates': models.Template.objects.all(),
                'form_template': form_template
            })


class TemplatePrintView(DetailView):
    template_name = 'crm_accounting/template_print_view.html'
    model = models.Invoice

    def get_context_data(self, **kwargs):
        context = dict()
        invoice = self.get_object()
        context['object'] = invoice
        context['templates'] = forms.TemplatePrint(initial={'template': models.Template.objects.filter(default=True).first()})
        return context

    def post(self, request, *args, **kwargs):
        template_id = request.POST.get('template')
        invoice = self.get_object()
        if 'action_download' in request.POST.keys():
            return print_invoice(invoice, template_id)
        if 'action_send_email' in request.POST.keys():
            if print_invoice(invoice, template_id, email=True):
                messages.success(request, 'Письмо успешно отправлено.')
            else:
                messages.error(request, 'У владельца не указана почта')
            return HttpResponseRedirect(
                reverse_lazy('crm_accounting:invoice_list'))



class SectionAjaxView(View):
    @staticmethod
    def get(request):
        house_id = request.GET.get('house_id')
        sections = list()
        if house_id:
            query_sections = Section.objects.filter(house_id=house_id)
            for section in query_sections:
                instance = {'section_id': section.id,
                            'section_title': section.title}
                sections.append(instance)
        return JsonResponse({'data': sections})


class FlatAjaxList(View):
    @staticmethod
    def get(request):
        section_id = request.GET.get('section_id')
        house_id = request.GET.get('house_id')
        flats = list()
        if section_id:
            query_flats = Flat.objects.filter(section_id=section_id)
            for flat in query_flats:
                instance = {'flat_id': flat.id,
                            'flat_title': flat.number}
                flats.append(instance)
        elif house_id:
            query_flats = Flat.objects.filter(house_id=house_id)
            for flat in query_flats:
                instance = {'flat_id': flat.id,
                            'flat_title': flat.number}
                flats.append(instance)
        return JsonResponse({'flats': flats})


class FlatAjaxInfo(View):
    @staticmethod
    def get(request):
        flat_id = request.GET.get('flat_id')
        data = dict()
        flat = get_object_or_404(Flat, id=flat_id)
        data['owner'] = {'owner': f'{flat.owner.first_name} '
                                  f'{flat.owner.last_name}',
                         'phone': flat.owner.phone,
                         'id': flat.owner.id}
        if flat.personal_account:
            data['personal_account'] = flat.personal_account.account_number

        return JsonResponse({'data': data})


def tariff_ajax_info(request):
    tariff_id = request.GET.get('tariff')
    services = TariffService.objects.filter(tariff_id=tariff_id).\
        select_related('service__unit')
    services_list = list()
    for service in services:
        instance = {
            'service_id': service.service.id,
            'unit_id': service.service.unit.id,
            'price': service.price
        }
        services_list.append(instance)
    return JsonResponse({'services': services_list})
